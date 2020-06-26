'''analysis to get most infos for the RQs in the paper
includes code for preprocessing commit diffs and 
calculating fix complexity'''
# import statement
import pymysql
import sys
import numpy as np
import datetime
import re
import subprocess
import shlex
import os
from openpyxl import Workbook
import dateutil.parser as dp

# save the main path from where this script is running
mainpath = os.getcwd()

# read the project name
project = str(sys.argv[1])
path = "/Users/nasifimtiaz/Desktop/new_data/"+sys.argv[2]
start=sys.argv[3]
end=sys.argv[4]


# get alerts with no event history and make a temporary table for them 
def fixed_alerts_with_history():
        with connection.cursor() as cursor:
                query = '''create temporary table alerts_with_no_history
                        select alert_id
                        from
                        (select a.idalerts as id, count(*) as c
                        from alerts a
                        join occurrences o
                        on a.idalerts=o.alert_id
                        where a.stream="''' + project + \
                        '''" group by a.idalerts) as event_count
                        join occurrences o
                        on event_count.id=o.alert_id
                        where event_count.c=1 
                        and o.line_number=1 '''
                cursor.execute(query)

                query="select * from alerts where stream='"+project+"' and is_invalid=0 and status='Fixed' and idalerts not in (select alert_id from alerts_with_no_history);"
                cursor.execute(query)
                return cursor.fetchall()

def all_file_fix_activity():
        query=query="select * from alerts where is_invalid=0 ans status='Fixed' and stream='"+project+"';"
        all_alerts=execute(query)
        for alert in all_alerts:
                aid=str(alert['idalerts'])
                cid=str(alert['cid'])
                print (aid)
                query="select * from occurrences where alert_id="+aid
                occurrences=execute(query)
                locations={}
                for o in occurrences:
                        fid=o["file_id"]
                        if fid is None:
                                continue
                        line=o["line_number"]
                        if fid not in locations.keys():
                                locations[fid]=[]
                        locations[fid].append(line)

                actionable=False
                commits=[]

                last_snapshot=str(alert["last_snapshot_id"])
                # get last detected dates
                query="select date from snapshots where idsnapshots="+last_snapshot+" and stream='" +project +"';"
                last_detected_date=execute(query)[0]["date"]
                last_detected_date=last_detected_date.strftime("%Y-%m-%d") + " 00:00:00"
                query="select date from snapshots where last_snapshot="+last_snapshot+" and stream='" +project +"';"
                first_not_detected_anymore_date=execute(query)[0]["date"]
                first_not_detected_anymore_date=first_not_detected_anymore_date.strftime("%Y-%m-%d") + " 00:00:00"

                for fid in locations.keys():
                        # look at if there's a commit within above two dates
                        print("file id is: ",fid)
                        query='''select * from filecommits f join commits c on f.commit_id=c.idcommits where
                                f.file_id= ''' + str(fid) + ''' and c.commit_date >="''' +last_detected_date+ \
                                '''" and c.commit_date <="''' +first_not_detected_anymore_date +'";'
                        results=execute(query)
                        if len(results)>0:
                                actionable=True
                                for item in results:
                                        commits.append(item)
                print(aid,commits)

def developer_fix_report():
        query="select count(*) as c from alerts where is_invalid=0 and stream='"+project+"';"
        total=execute(query)[0]['c']

        query='''select count(*) as c from actionability ac 
                join alerts a 
                on ac.alert_id=a.idalerts
                where ac.actionable=1 
                and a.stream="'''+project+'''";'''
        fix=execute(query)[0]['c']

        query='''select count(*) as c from alert_commit_tracking ac 
                join alerts a 
                on ac.alert_id=a.idalerts
                where ac.fix_commit_id is not null
                and a.stream="'''+project+'''";'''
        fix_commit=execute(query)[0]['c']

        unactionable=total-fix

        f.write("\ndeveloper fix :"+str(fix)+" "+str((float(fix)/total)*100)+'\n')
        f.write("tracked fix commit for :"+str(fix_commit)+" "+str((float(fix_commit)/total)*100)+'\n')
        f.write("unactionable :"+str(unactionable)+" "+str((float(unactionable)/total)*100)+'\n')

def actionability_and_lifespan_report():
        query="select count(*) as c from alerts where is_invalid=0 and stream='"+project+"';"
        total_alerts=execute(query)[0]['c']
        # need to subtract new alerts that are alive for less the median lifespan of actionable alerts
        
        query='''select count(*) as c from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                where ac.actionability = 1
                and a.is_invalid=0
                and a.stream="''' + project + '"'
        actionable_count=execute(query)[0]['c']
        
        query='''select datediff(s.date,a.first_detected) as diff from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability = 1
                and a.is_invalid=0
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        actionable_lifespan=np.median(temp)

        query='''select count(*) as c
                from alerts 
                where stream="'''+project+'''"
                and status='New'
                and is_invalid=0
                and datediff(
                (select date
                from snapshots
                where stream="'''+project+'''"
                order by idsnapshots desc
                limit 1),
                first_detected
                ) <=''' +str(actionable_lifespan)
        too_new_to_count=execute(query)[0]['c']
        total_alerts-=too_new_to_count



        f.write("actionable bugs are: "+str(actionable_count)+'\n')
        f.write("median lifspan of actionable alerts are : "+str(actionable_lifespan)+'\n')
        f.write("actionablity rate: "+str((float(actionable_count)/total_alerts)*100)+'\n')

        query='''select datediff(s.date,a.first_detected) as diff from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability = 1
                and a.is_invalid=0
                and a.classification="Bug"
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        bug_lifespan=np.median(temp)
        f.write("median lifspan of  marked bug are : "+str(bug_lifespan)+'\n')

       # 3 is for other-file alerts  
        query='''select datediff(s.date,a.first_detected) as diff 
                from alerts a 
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                and a.is_invalid=3
                and a.status='Fixed'
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        other_file_lifespan=np.median(temp)
        f.write("median lifspan of other-file alerts are : "+str(other_file_lifespan)+'\n')

        query='''select datediff(s.date,a.first_detected) as diff from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability = 0
                and a.is_invalid=0
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        unactionable_lifespan=np.median(temp)
        f.write("median lifspan of unactionable alerts are : "+str(unactionable_lifespan)+'\n')

def manual_validation_file():
        os.chdir(mainpath)

        wb=Workbook()
        ws1=wb.create_sheet("Fix commits",0)
        query='''select a.idalerts,a.cid, b.*,f.filepath_on_coverity, ac.*, datediff(s.date,a.first_detected) as diff
                from actionability ac
                join alerts a
                on a.idalerts=ac.alert_id 
                join files f
                on f.idfiles=a.file_id
                join bug_types b
                on b.idbug_types=a.bug_type
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability=1
                and (ac.single_fix_commit is not null or ac.fix_commits is not null)
                and a.stream="'''+project+'''"
                order by rand()
                limit 50'''
        results=execute(query)

        row=2
        for item in results:
                # get commit hashes
                commits=[]
                if item['single_fix_commit']!=None:
                        query="select * from commits where idcommits="+str(item['single_fix_commit'])
                        commits.append(execute(query)[0])
                else:
                        temp=str(item['fix_commits'])
                        temp=temp.split(',')
                        for t in temp:
                                query="select * from commits where idcommits="+str(t)
                                commits.append(execute(query)[0])
                for c in commits:    
                        ws1['A'+str(row)]=item['idalerts']
                        ws1['B'+str(row)]=item['cid']
                        ws1['C'+str(row)]=item['type']
                        ws1['D'+str(row)]=item['impact']
                        ws1['E'+str(row)]=item['filepath_on_coverity']
                        # look at lifespan
                        ws1['F'+str(row)]=item['diff']
                        # complexity I can check on the commit diffs itself
                        ws1['G'+str(row)]=c['sha']
                        # if len(commits)>1:
                        #         ws1['H'+str(row)]="\\b"+c['message'].encode('ascii','ignore')+"\\b0"
                        # else:
                        ws1['H'+str(row)]=c['message'].encode('ascii','ignore').decode()
                        row+=1

        ws2=wb.create_sheet("Unactionable Alerts",1)

        query='''select a.idalerts,a.cid, b.*, f.filepath_on_coverity, s.date as lastdate, ac.*
                from actionability ac
                join alerts a
                on a.idalerts=ac.alert_id 
                join files f
                on f.idfiles=a.file_id
                join bug_types b
                on b.idbug_types=a.bug_type
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability=0
                and (ac.file_deleted is null and ac.file_renamed is null)
                and a.stream="'''+project+'''"
                order by rand()
                limit 50'''
        results=execute(query)

        row=2
        for item in results:
                ws2['A'+str(row)]=item['idalerts']
                ws2['B'+str(row)]=item['cid']
                ws2['C'+str(row)]=item['type']
                ws2['D'+str(row)]=item['impact']
                ws2['E'+str(row)]= item['lastdate']
                ws2['F'+str(row)]=item['filepath_on_coverity']
                ws2['G'+str(row)]=item['file_deleted']
                ws2['H'+str(row)]=item['file_renamed']
                ws2['I'+str(row)]=item['suppression']
                ws2['J'+str(row)]='*'
                row+=1
        
        ws3=wb.create_sheet("Single Fix Commits",2)
        query='''select a.idalerts,a.cid, b.*, f.filepath_on_coverity, ac.*, datediff(s.date,a.first_detected) as diff
                from actionability ac
                join alerts a
                on a.idalerts=ac.alert_id 
                join files f
                on f.idfiles=a.file_id
                join bug_types b
                on b.idbug_types=a.bug_type
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability=1
                and ac.single_fix_commit is not null
                and a.stream="'''+project+'''"
                order by rand()
                limit 50'''
        results=execute(query)

        row=2
        for item in results:
                # get commit hashes
                commits=[]
                if item['single_fix_commit']!=None:
                        query="select * from commits c\
                                join merge_date m on c.idcommits=m.merge_date \
                                where idcommits="+str(item['single_fix_commit'])
                        commits.append(execute(query)[0])
                for c in commits:   
                        ws3['A'+str(row)]=item['idalerts']
                        ws3['B'+str(row)]=item['cid']
                        ws3['C'+str(row)]=item['type']
                        ws3['D'+str(row)]=item['impact']
                        ws3['E'+str(row)]=item['filepath_on_coverity']
                        # look at lifespan
                        ws3['F'+str(row)]=item['diff']
                        # complexity I can check on the commit diffs itself
                        ws3['G'+str(row)]=c['sha']
                        ws3['H'+str(row)]=c['message'].encode('ascii',errors='ignore').decode()
                        ws3['I'+str(row)]=c['merge_date']
                        row+=1

        wb.save('Project_'+project+'.xlsx')

def file_renaming_info():
        #Need to change this
        # we search this only for fixed alerts as
        # alerts are supposed to be eliminated 
        # when the file is renamed (deleted)
        query='''select count(*) as c
                from alerts
                where stream= "''' +project+'''"
                and idalerts in
                (select alert_id from actionability 
                where file_renamed="yes")''' 
        c=execute(query)[0]['c']
        f.write('the number of alerts that are invalidated due to file renaming is: '+str(c)+'\n')

        query='''select count(*) as c
                from alerts
                where stream= "''' +project+'''"
                and idalerts in
                (select alert_id from actionability 
                where file_renamed="yes"
                and transfered_alert_id is not null)''' 
        c=execute(query)[0]['c']
        f.write('the number of alerts that are transfered due to file renaming is: '+str(c)+'\n')



'''Code starts here'''
def patch_complexity():
        # revoking this function
        query='''select count(*) as c
                from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                where a.stream="'''+project +'''"
                and a.is_invalid=0
                and ac.single_fix_commit is not null'''
        fix_commit_tracked=execute(query)[0]['c']
        f.write("fix commit tracked for alerts: "+str(fix_commit_tracked)+'\n')

        

        query='''select fc.*
                from alerts a
                join actionability ac
                on a.idalerts=ac.alert_id
                join fix_complexity fc
                on fc.alert_id=ac.alert_id
                and fc.commit_id=ac.single_fix_commit
                where a.stream="'''+project+'''"
                and a.status='Fixed'
                and a.is_invalid=0
                and ac.single_fix_commit is not null;'''
        results=execute(query)
        files=[]
        net_loc=[]
        net_logical=[]
        in_loc=[]
        in_logical=[]
        for item in results:
                file_count=float(item['file_count'])/float(item['total_fixed_alerts'])
                loc_change=float(item['net_loc_change'])/float(item['total_fixed_alerts'])
                logical_change=float(item['net_logical_change'])/(item['total_fixed_alerts'])
                files.append(file_count)
                net_loc.append(loc_change)
                net_logical.append(logical_change)
                loc_change=float(item['infile_loc_change'])/float(item['infile_fixed_alerts'])
                logical_change=float(item['infile_logical_change'])/(item['infile_fixed_alerts'])
                in_loc.append(loc_change)
                in_logical.append(logical_change)
        f.write("median affected files: "+str(np.median(files))+'\n')
        f.write("median net LOC change: "+str(np.median(net_loc))+'\n')
        f.write("median net logical change: "+str(np.median(net_logical))+'\n')
        f.write("median infile LOC change: "+str(np.median(in_loc))+'\n')
        f.write("median infile logical change: "+str(np.median(in_logical))+'\n')
        # query='''select * from
        # (select ac.single_fix_commit, a.file_id, count(*)  as c
        # from actionability ac
        # join alerts a
        # on ac.alert_id=a.idalerts 
        # where a.stream="'''+project+'''"
        # and ac.single_fix_commit is not null
        # and a.is_invalid=0
        # group by ac.single_fix_commit,a.file_id) as fix
        # join filecommits fc
        # on fix.single_fix_commit=fc.commit_id'''
        # results=execute(query)
        # loc=[]
        # for item in results:
        #         if item['lines_added']==None or item['lines_removed']==None:
        #                 continue
        #         c=item['c']
        #         l=float(item['lines_added']+item['lines_removed'])
        #         loc.append(l/c)
        # loc.sort()
        # f.write("median in-file LOC change: "+str(np.median(loc)))

def methodology_infos():
        get_general_report()
        # need to give info on data cleaning
        file_renaming_info()








                    
if __name__ == "__main__":
    # # reporting functions
    methodology_infos()
    actionability_and_lifespan_report()
    
    update_fix_commit_infos()
    patch_complexity()
    manual_validation_file()

    f.close()
