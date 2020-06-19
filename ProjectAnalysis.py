'''analysis to get most infos for the RQs in the paper
includes code for preprocessing commit diffs and 
calculating fix complexity'''
# import statement
import pymysql
import sys
import numpy as np
import datetime
import re
import subprocess
import shlex
import os
from openpyxl import Workbook
import dateutil.parser as dp

# save the main path from where this script is running
mainpath = os.getcwd()

# read the project name
project = str(sys.argv[1])
path = "/Users/nasifimtiaz/Desktop/new_data/"+sys.argv[2]
start=sys.argv[3]
end=sys.argv[4]


# get alerts with no event history and make a temporary table for them 
def fixed_alerts_with_history():
        with connection.cursor() as cursor:
                query = '''create temporary table alerts_with_no_history
                        select alert_id
                        from
                        (select a.idalerts as id, count(*) as c
                        from alerts a
                        join occurrences o
                        on a.idalerts=o.alert_id
                        where a.stream="''' + project + \
                        '''" group by a.idalerts) as event_count
                        join occurrences o
                        on event_count.id=o.alert_id
                        where event_count.c=1 
                        and o.line_number=1 '''
                cursor.execute(query)

                query="select * from alerts where stream='"+project+"' and is_invalid=0 and status='Fixed' and idalerts not in (select alert_id from alerts_with_no_history);"
                cursor.execute(query)
                return cursor.fetchall()

def all_file_fix_activity():
        query=query="select * from alerts where is_invalid=0 ans status='Fixed' and stream='"+project+"';"
        all_alerts=execute(query)
        for alert in all_alerts:
                aid=str(alert['idalerts'])
                cid=str(alert['cid'])
                print (aid)
                query="select * from occurrences where alert_id="+aid
                occurrences=execute(query)
                locations={}
                for o in occurrences:
                        fid=o["file_id"]
                        if fid is None:
                                continue
                        line=o["line_number"]
                        if fid not in locations.keys():
                                locations[fid]=[]
                        locations[fid].append(line)

                actionable=False
                commits=[]

                last_snapshot=str(alert["last_snapshot_id"])
                # get last detected dates
                query="select date from snapshots where idsnapshots="+last_snapshot+" and stream='" +project +"';"
                last_detected_date=execute(query)[0]["date"]
                last_detected_date=last_detected_date.strftime("%Y-%m-%d") + " 00:00:00"
                query="select date from snapshots where last_snapshot="+last_snapshot+" and stream='" +project +"';"
                first_not_detected_anymore_date=execute(query)[0]["date"]
                first_not_detected_anymore_date=first_not_detected_anymore_date.strftime("%Y-%m-%d") + " 00:00:00"

                for fid in locations.keys():
                        # look at if there's a commit within above two dates
                        print("file id is: ",fid)
                        query='''select * from filecommits f join commits c on f.commit_id=c.idcommits where
                                f.file_id= ''' + str(fid) + ''' and c.commit_date >="''' +last_detected_date+ \
                                '''" and c.commit_date <="''' +first_not_detected_anymore_date +'";'
                        results=execute(query)
                        if len(results)>0:
                                actionable=True
                                for item in results:
                                        commits.append(item)
                print(aid,commits)

def developer_fix_report():
        query="select count(*) as c from alerts where is_invalid=0 and stream='"+project+"';"
        total=execute(query)[0]['c']

        query='''select count(*) as c from actionability ac 
                join alerts a 
                on ac.alert_id=a.idalerts
                where ac.actionable=1 
                and a.stream="'''+project+'''";'''
        fix=execute(query)[0]['c']

        query='''select count(*) as c from alert_commit_tracking ac 
                join alerts a 
                on ac.alert_id=a.idalerts
                where ac.fix_commit_id is not null
                and a.stream="'''+project+'''";'''
        fix_commit=execute(query)[0]['c']

        unactionable=total-fix

        f.write("\ndeveloper fix :"+str(fix)+" "+str((float(fix)/total)*100)+'\n')
        f.write("tracked fix commit for :"+str(fix_commit)+" "+str((float(fix_commit)/total)*100)+'\n')
        f.write("unactionable :"+str(unactionable)+" "+str((float(unactionable)/total)*100)+'\n')

def actionability_and_lifespan_report():
        query="select count(*) as c from alerts where is_invalid=0 and stream='"+project+"';"
        total_alerts=execute(query)[0]['c']
        # need to subtract new alerts that are alive for less the median lifespan of actionable alerts
        
        query='''select count(*) as c from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                where ac.actionability = 1
                and a.is_invalid=0
                and a.stream="''' + project + '"'
        actionable_count=execute(query)[0]['c']
        
        query='''select datediff(s.date,a.first_detected) as diff from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability = 1
                and a.is_invalid=0
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        actionable_lifespan=np.median(temp)

        query='''select count(*) as c
                from alerts 
                where stream="'''+project+'''"
                and status='New'
                and is_invalid=0
                and datediff(
                (select date
                from snapshots
                where stream="'''+project+'''"
                order by idsnapshots desc
                limit 1),
                first_detected
                ) <=''' +str(actionable_lifespan)
        too_new_to_count=execute(query)[0]['c']
        total_alerts-=too_new_to_count



        f.write("actionable bugs are: "+str(actionable_count)+'\n')
        f.write("median lifspan of actionable alerts are : "+str(actionable_lifespan)+'\n')
        f.write("actionablity rate: "+str((float(actionable_count)/total_alerts)*100)+'\n')

        query='''select datediff(s.date,a.first_detected) as diff from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability = 1
                and a.is_invalid=0
                and a.classification="Bug"
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        bug_lifespan=np.median(temp)
        f.write("median lifspan of  marked bug are : "+str(bug_lifespan)+'\n')

       # 3 is for other-file alerts  
        query='''select datediff(s.date,a.first_detected) as diff 
                from alerts a 
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                and a.is_invalid=3
                and a.status='Fixed'
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        other_file_lifespan=np.median(temp)
        f.write("median lifspan of other-file alerts are : "+str(other_file_lifespan)+'\n')

        query='''select datediff(s.date,a.first_detected) as diff from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability = 0
                and a.is_invalid=0
                and a.stream="''' + project + '"'
        results=execute(query)
        temp=[]
        for item in results:
                temp.append(item['diff'])
        
        unactionable_lifespan=np.median(temp)
        f.write("median lifspan of unactionable alerts are : "+str(unactionable_lifespan)+'\n')

def manual_validation_file():
        os.chdir(mainpath)

        wb=Workbook()
        ws1=wb.create_sheet("Fix commits",0)
        query='''select a.idalerts,a.cid, b.*,f.filepath_on_coverity, ac.*, datediff(s.date,a.first_detected) as diff
                from actionability ac
                join alerts a
                on a.idalerts=ac.alert_id 
                join files f
                on f.idfiles=a.file_id
                join bug_types b
                on b.idbug_types=a.bug_type
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability=1
                and (ac.single_fix_commit is not null or ac.fix_commits is not null)
                and a.stream="'''+project+'''"
                order by rand()
                limit 50'''
        results=execute(query)

        row=2
        for item in results:
                # get commit hashes
                commits=[]
                if item['single_fix_commit']!=None:
                        query="select * from commits where idcommits="+str(item['single_fix_commit'])
                        commits.append(execute(query)[0])
                else:
                        temp=str(item['fix_commits'])
                        temp=temp.split(',')
                        for t in temp:
                                query="select * from commits where idcommits="+str(t)
                                commits.append(execute(query)[0])
                for c in commits:    
                        ws1['A'+str(row)]=item['idalerts']
                        ws1['B'+str(row)]=item['cid']
                        ws1['C'+str(row)]=item['type']
                        ws1['D'+str(row)]=item['impact']
                        ws1['E'+str(row)]=item['filepath_on_coverity']
                        # look at lifespan
                        ws1['F'+str(row)]=item['diff']
                        # complexity I can check on the commit diffs itself
                        ws1['G'+str(row)]=c['sha']
                        # if len(commits)>1:
                        #         ws1['H'+str(row)]="\\b"+c['message'].encode('ascii','ignore')+"\\b0"
                        # else:
                        ws1['H'+str(row)]=c['message'].encode('ascii','ignore').decode()
                        row+=1

        ws2=wb.create_sheet("Unactionable Alerts",1)

        query='''select a.idalerts,a.cid, b.*, f.filepath_on_coverity, s.date as lastdate, ac.*
                from actionability ac
                join alerts a
                on a.idalerts=ac.alert_id 
                join files f
                on f.idfiles=a.file_id
                join bug_types b
                on b.idbug_types=a.bug_type
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability=0
                and (ac.file_deleted is null and ac.file_renamed is null)
                and a.stream="'''+project+'''"
                order by rand()
                limit 50'''
        results=execute(query)

        row=2
        for item in results:
                ws2['A'+str(row)]=item['idalerts']
                ws2['B'+str(row)]=item['cid']
                ws2['C'+str(row)]=item['type']
                ws2['D'+str(row)]=item['impact']
                ws2['E'+str(row)]= item['lastdate']
                ws2['F'+str(row)]=item['filepath_on_coverity']
                ws2['G'+str(row)]=item['file_deleted']
                ws2['H'+str(row)]=item['file_renamed']
                ws2['I'+str(row)]=item['suppression']
                ws2['J'+str(row)]='*'
                row+=1
        
        ws3=wb.create_sheet("Single Fix Commits",2)
        query='''select a.idalerts,a.cid, b.*, f.filepath_on_coverity, ac.*, datediff(s.date,a.first_detected) as diff
                from actionability ac
                join alerts a
                on a.idalerts=ac.alert_id 
                join files f
                on f.idfiles=a.file_id
                join bug_types b
                on b.idbug_types=a.bug_type
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability=1
                and ac.single_fix_commit is not null
                and a.stream="'''+project+'''"
                order by rand()
                limit 50'''
        results=execute(query)

        row=2
        for item in results:
                # get commit hashes
                commits=[]
                if item['single_fix_commit']!=None:
                        query="select * from commits c\
                                join merge_date m on c.idcommits=m.merge_date \
                                where idcommits="+str(item['single_fix_commit'])
                        commits.append(execute(query)[0])
                for c in commits:   
                        ws3['A'+str(row)]=item['idalerts']
                        ws3['B'+str(row)]=item['cid']
                        ws3['C'+str(row)]=item['type']
                        ws3['D'+str(row)]=item['impact']
                        ws3['E'+str(row)]=item['filepath_on_coverity']
                        # look at lifespan
                        ws3['F'+str(row)]=item['diff']
                        # complexity I can check on the commit diffs itself
                        ws3['G'+str(row)]=c['sha']
                        ws3['H'+str(row)]=c['message'].encode('ascii',errors='ignore').decode()
                        ws3['I'+str(row)]=c['merge_date']
                        row+=1

        wb.save('Project_'+project+'.xlsx')

def file_renaming_info():
        #Need to change this
        # we search this only for fixed alerts as
        # alerts are supposed to be eliminated 
        # when the file is renamed (deleted)
        query='''select count(*) as c
                from alerts
                where stream= "''' +project+'''"
                and idalerts in
                (select alert_id from actionability 
                where file_renamed="yes")''' 
        c=execute(query)[0]['c']
        f.write('the number of alerts that are invalidated due to file renaming is: '+str(c)+'\n')

        query='''select count(*) as c
                from alerts
                where stream= "''' +project+'''"
                and idalerts in
                (select alert_id from actionability 
                where file_renamed="yes"
                and transfered_alert_id is not null)''' 
        c=execute(query)[0]['c']
        f.write('the number of alerts that are transfered due to file renaming is: '+str(c)+'\n')



'''Code starts here'''
def patch_complexity():
        # revoking this function
        query='''select count(*) as c
                from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                where a.stream="'''+project +'''"
                and a.is_invalid=0
                and ac.single_fix_commit is not null'''
        fix_commit_tracked=execute(query)[0]['c']
        f.write("fix commit tracked for alerts: "+str(fix_commit_tracked)+'\n')

        

        query='''select fc.*
                from alerts a
                join actionability ac
                on a.idalerts=ac.alert_id
                join fix_complexity fc
                on fc.alert_id=ac.alert_id
                and fc.commit_id=ac.single_fix_commit
                where a.stream="'''+project+'''"
                and a.status='Fixed'
                and a.is_invalid=0
                and ac.single_fix_commit is not null;'''
        results=execute(query)
        files=[]
        net_loc=[]
        net_logical=[]
        in_loc=[]
        in_logical=[]
        for item in results:
                file_count=float(item['file_count'])/float(item['total_fixed_alerts'])
                loc_change=float(item['net_loc_change'])/float(item['total_fixed_alerts'])
                logical_change=float(item['net_logical_change'])/(item['total_fixed_alerts'])
                files.append(file_count)
                net_loc.append(loc_change)
                net_logical.append(logical_change)
                loc_change=float(item['infile_loc_change'])/float(item['infile_fixed_alerts'])
                logical_change=float(item['infile_logical_change'])/(item['infile_fixed_alerts'])
                in_loc.append(loc_change)
                in_logical.append(logical_change)
        f.write("median affected files: "+str(np.median(files))+'\n')
        f.write("median net LOC change: "+str(np.median(net_loc))+'\n')
        f.write("median net logical change: "+str(np.median(net_logical))+'\n')
        f.write("median infile LOC change: "+str(np.median(in_loc))+'\n')
        f.write("median infile logical change: "+str(np.median(in_logical))+'\n')
        # query='''select * from
        # (select ac.single_fix_commit, a.file_id, count(*)  as c
        # from actionability ac
        # join alerts a
        # on ac.alert_id=a.idalerts 
        # where a.stream="'''+project+'''"
        # and ac.single_fix_commit is not null
        # and a.is_invalid=0
        # group by ac.single_fix_commit,a.file_id) as fix
        # join filecommits fc
        # on fix.single_fix_commit=fc.commit_id'''
        # results=execute(query)
        # loc=[]
        # for item in results:
        #         if item['lines_added']==None or item['lines_removed']==None:
        #                 continue
        #         c=item['c']
        #         l=float(item['lines_added']+item['lines_removed'])
        #         loc.append(l/c)
        # loc.sort()
        # f.write("median in-file LOC change: "+str(np.median(loc)))

def methodology_infos():
        get_general_report()
        # need to give info on data cleaning
        file_renaming_info()

def update_fix_commit_infos():
        query='''select distinct ac.single_fix_commit, c.*,f.*,a.idalerts from actionability ac
                join alerts a
                on ac.alert_id=a.idalerts
                join commits c
                on c.idcommits = ac.single_fix_commit
                join files f
                on f.idfiles=a.file_id
                where ac.single_fix_commit is not null
                and ac.actionability=1
                and a.is_invalid=0
                and a.status='Fixed'
                and a.stream ="'''+project+'"'
        results=execute(query)

        for item in results:
                sha=item['sha']
                cid=item['idcommits']
                filepath=item['filepath_on_coverity'][1:]
                fid=item['idfiles']
                aid=item['idalerts']
                arguments=process_commit(sha,filepath)

                query='select count(*) as c from actionability where single_fix_commit='+str(cid)
                total_fixed_alerts=execute(query)[0]['c']
                arguments.append(total_fixed_alerts)

                query='''select count(*) as c from actionability ac
                        join alerts a
                        on a.idalerts = ac.alert_id
                        where ac.single_fix_commit='''+str(cid)+'''
                        and a.file_id='''+str(fid)
                infile_fixed_alerts=execute(query)[0]['c']
                arguments.append(infile_fixed_alerts)

                query="insert into fix_complexity values("+str(cid)+","+str(aid)+","
                for idx, arg in enumerate(arguments):
                    # value cleaning
                    arg=str(arg) #if not string
                    arg=arg.strip() #if any whitespace ahead or trailing
                    # remove illegal character
                    arg=arg.replace('"',"'")

                    if is_number(arg) or arg=="null":
                        query+=arg
                    else:
                        query+='"'+arg+'"'
                    if idx<len(arguments)-1:
                        query+=","
                query+=");"
                with connection.cursor() as cursor:
                    try:
                        cursor.execute(query)
                    except Exception as e:
                        print(e,query)
        # for item in results:
        #         d={}
        #         sha=item['sha']
        #         cid=item['idcommits']
        #         filepath=item['filepath_on_coverity'][1:]
        #         fid=item['idfiles']
        #         lines=subprocess.check_output(
        #                 shlex.split('git show --stat '+sha),
        #                 encoding="437"
        #         ).split('\n')
        #         line=lines[-2].strip()
        #         temp=line.split(',')
        #         info=[]
        #         query=''
        #         if len(temp)==3:
        #                 for t in temp:
        #                         t=t.strip()
        #                         info.append(t.split(' ')[0])
        #                 #update commit table
        #                 query='update commits set affected_files_count='+info[0]+',net_lines_added='+info[1]+',net_lines_removed='+info[2]+' where idcommits='+str(cid)
        #         elif len(temp)==2:
        #                 if 'insert' in line:
        #                         for t in temp:
        #                                 t=t.strip()
        #                                 info.append(t.split(' ')[0])
        #                         #update commit table
        #                         query='update commits set affected_files_count='+info[0]+',net_lines_added='+info[1]+',net_lines_removed=0 where idcommits='+str(cid)
        #                 elif 'delet' in line:
        #                         for t in temp:
        #                                 t=t.strip()
        #                                 info.append(t.split(' ')[0])
        #                         #update commit table
        #                         query='update commits set affected_files_count='+info[0]+',net_lines_added=0,net_lines_removed='+info[1]+' where idcommits='+str(cid)
        #         execute(query)  



def lifespan_vs_complexity():
        query='''select idalerts, datediff(s.date,a.first_detected) as diff from actionability ac
                join alerts a 
                on a.idalerts=ac.alert_id
                join snapshots s
                on a.last_snapshot_id=s.idsnapshots
                where ac.actionability = 1
                and a.stream="''' + project + '"'

def remove_blank_lines_and_comments(diff):
        diff=re.sub('/\\*(.|\n)*\\*/','',diff) # remmoving multiline comments
        lines=diff.split('\n')
        copy=[]
        for line in lines:
                line=re.sub('//(.*)','',line) #removing single line comments
                line=line.strip() #stripping trailing whitespaces
                if not (line=='' or line=='\n'):
                        # however changes will start with + or -
                        if line.startswith('+') or line.startswith('-'):
                                temp=line[1:] #take the actual line
                                temp=temp.strip() #stripping trailing whitespaces
                                if not (temp=='' or temp=='\n'):
                                        copy.append(line)
                        else:
                                copy.append(line)
        return '\n'.join(copy)
def process_commit(sha,filepath):
        affected_files=0
        net_loc_change=0
        infile_loc_change=0
        net_logical_change=0
        infile_logical_change=0
        full=subprocess.check_output(
                shlex.split("git show "+sha),
                encoding="437"
        )
        t=open('temp.txt','w')
        t.write(full)
        t.close()

        diffs=full.split('diff --git ')
        del diffs[0]
        affected_files=len(diffs)

        for diff in diffs:
                loc=0
                logical=0
                diff=re.sub('[@]{3,}','',diff)
                parts=diff.split('@@')
                del parts[0]
                ind = 1
                while ind < len(parts):
                        cur_diff=parts[ind]
                        cur_diff=remove_blank_lines_and_comments(cur_diff)
                        lines=cur_diff.split('\n')
                        flag=False #flag to keep track of logical changes
                        for line in lines:
                                if line.startswith('+') or line.startswith('-'):
                                        loc+=1
                                        if not flag:
                                                logical+=1
                                                flag=True
                                else:
                                        flag=False
                        ind+=2
                if filepath in diff:
                       infile_loc_change=loc
                       infile_logical_change=logical
                net_loc_change+=loc
                net_logical_change+=logical
        
        return [affected_files,net_loc_change,infile_loc_change,net_logical_change,infile_logical_change]

                        


        pass
if __name__ == "__main__":
    # # reporting functions
    methodology_infos()
    actionability_and_lifespan_report()
    
    update_fix_commit_infos()
    patch_complexity()
    manual_validation_file()

    f.close()
